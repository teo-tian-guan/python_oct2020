import openpyxl

workbook = openpyxl.load_workbook("bmi.xlsx")
sheet=workbook["Sheet1"]

max_row = sheet.max_row
max_column = sheet.max_column
students = []
#loop through every row
for i in range(2,max_row+1):
    
    #read cell
    name = sheet.cell(row=i, column=1).value
    weight = sheet.cell(row=i, column=2).value
    height = sheet.cell(row=i, column=3).value
    
    weight = float(weight)
    height = float(height)
    bmi = weight/(height * height)
    
    students.append([name, weight, height, bmi])
    
    #print(name + ":" + str(weight) + ":" + str(height))
    
import csv

writerFileHandle = open("student_bmi.csv", "w", newline='')
writer1 = csv.writer(writerFileHandle)
writer1.writerows(students)
writerFileHandle.close()
